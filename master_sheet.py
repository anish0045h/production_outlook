"""
master_sheet.py — Manages the master_payroll.xlsx file.

Creates the file + header row if it doesn't exist.
Appends records without duplicates (keyed on Employee ID + Month + Source File).
Applies formatting: header bold, alternating row colours, auto column widths.
"""

import os
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

HEADERS = [
    "Employee ID",
    "Employee Name",
    "Designation",
    "ASE Manager",
    "ASM Manager",
    "Month",
    "Total Extra KM",
    "Amount INR",
    "Source File",
    "Email Subject",
    "Sender Email",
    "Processed Date",
]

# Styling constants
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
ROW_FILL_ODD  = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="EBF3FB")   # light blue
TOTAL_FILL    = PatternFill("solid", fgColor="FFF2CC")   # light yellow

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SHEET_NAME = "Payroll Records"


class MasterSheet:

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb, self.ws = self._load_or_create()
        self._existing_keys = self._build_key_index()

    # ── Load / create ──────────────────────────────────────────────

    def _load_or_create(self):
        if os.path.exists(self.filepath):
            log.info(f"Loading existing master file: {self.filepath}")
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        else:
            log.info(f"Creating new master file: {self.filepath}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            self._write_header(ws)

        return wb, ws

    def _write_header(self, ws):
        ws.append(HEADERS)
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 30

    # ── Duplicate detection ────────────────────────────────────────

    def _build_key_index(self) -> set:
        """
        Build a set of (employee_id, month, source_file) tuples
        from existing rows so we don't add duplicates.
        """
        keys = set()
        emp_col    = HEADERS.index("Employee ID")
        month_col  = HEADERS.index("Month")
        source_col = HEADERS.index("Source File")

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row[emp_col] is not None:
                keys.add((
                    str(row[emp_col]).strip(),
                    str(row[month_col]).strip(),
                    str(row[source_col]).strip(),
                ))
        return keys

    # ── Append ─────────────────────────────────────────────────────

    def append_records(self, records: list[dict]) -> int:
        """Append non-duplicate records. Returns count of rows added."""
        added = 0
        for rec in records:
            key = (
                str(rec.get("Employee ID", "")).strip(),
                str(rec.get("Month", "")).strip(),
                str(rec.get("Source File", "")).strip(),
            )
            if key in self._existing_keys:
                log.debug(f"Duplicate skipped: {key}")
                continue

            row_values = [rec.get(h, "") for h in HEADERS]
            self.ws.append(row_values)
            self._style_data_row(self.ws.max_row)
            self._existing_keys.add(key)
            added += 1

        return added

    def _style_data_row(self, row_num: int):
        fill = ROW_FILL_EVEN if row_num % 2 == 0 else ROW_FILL_ODD
        for col_idx, cell in enumerate(self.ws[row_num], start=1):
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            # Right-align numeric columns
            header = HEADERS[col_idx - 1]
            if header in ("Employee ID", "Total Extra KM", "Amount INR"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if header == "Amount INR":
                    cell.number_format = '₹#,##0'
                if header == "Total Extra KM":
                    cell.number_format = '#,##0'

    # ── Save ───────────────────────────────────────────────────────

    def save(self):
        self._add_totals_row()
        self._auto_fit_columns()
        self._freeze_header()
        self.wb.save(self.filepath)
        log.info(f"Master file saved: {self.filepath}  ({self.ws.max_row - 2} data rows)")

    def _add_totals_row(self):
        """Add / refresh the totals row at the bottom."""
        # Remove old totals row if it exists
        last_row = self.ws.max_row
        if last_row > 1:
            last_values = [self.ws.cell(last_row, c).value for c in range(1, 4)]
            if last_values[0] == "TOTAL":
                self.ws.delete_rows(last_row)

        total_row = self.ws.max_row + 1
        data_rows = total_row - 2  # excluding header

        if data_rows <= 0:
            return

        km_col     = get_column_letter(HEADERS.index("Total Extra KM") + 1)
        amount_col = get_column_letter(HEADERS.index("Amount INR") + 1)

        self.ws.cell(total_row, 1).value = "TOTAL"
        self.ws.cell(total_row, HEADERS.index("Total Extra KM") + 1).value = (
            f"=SUM({km_col}2:{km_col}{total_row-1})"
        )
        self.ws.cell(total_row, HEADERS.index("Amount INR") + 1).value = (
            f"=SUM({amount_col}2:{amount_col}{total_row-1})"
        )

        for col_idx in range(1, len(HEADERS) + 1):
            cell = self.ws.cell(total_row, col_idx)
            cell.fill   = TOTAL_FILL
            cell.font   = Font(bold=True, size=11)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right", vertical="center")

        self.ws.cell(total_row, 1).alignment = Alignment(horizontal="left", vertical="center")

    def _auto_fit_columns(self):
        """Set column widths based on content."""
        col_widths = {h: len(h) for h in HEADERS}
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            for col_idx, val in enumerate(row):
                if val is not None:
                    col_widths[HEADERS[col_idx]] = max(
                        col_widths[HEADERS[col_idx]], min(len(str(val)), 40)
                    )
        for col_idx, header in enumerate(HEADERS, start=1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = (
                col_widths[header] + 4
            )

    def _freeze_header(self):
        """Freeze the header row so it stays visible while scrolling."""
        self.ws.freeze_panes = "A2"
