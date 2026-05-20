"""
excel_parser.py — Parses the CSO travel Excel file and returns
clean payroll records ready to append to the master sheet.

Expected sheet structure (sheet: "Present absent"):
  Row 1-4 : Date headers (Row 4 contains actual datetime objects)
  Row 5   : Column headers
  Row 6+  : One employee per row

Actual column layout confirmed from file inspection (0-based):
  0  = Employee ID
  1  = Employee Name
  2  = Designation
  3  = ASE Manager
  4  = ASM Manager
  5-34 = Daily KM (30 days — col 5 to col 34 inclusive)
  35 = Exception Kms in the Month (to be approved by SM)
  36 = SM Approval   ← "Yes" or blank

Period detection (NEW):
  Month tag is now extracted from the actual datetime values in Row 4
  (first date → last date), giving reliable period tagging independent
  of filename or email subject wording.
  Falls back to the original filename/subject text method if Row 4
  contains no datetime objects.
"""

import io
import re
import logging
from datetime import datetime, date
import openpyxl

log = logging.getLogger(__name__)

# ── Column indices (0-based) ───────────────────────────────────────────────────
COL_EMP_ID    = 0   # A
COL_EMP_NAME  = 1   # B
COL_DESIG     = 2   # C
COL_ASE       = 3   # D
COL_ASM       = 4   # E
COL_KM_START  = 5   # F  — first daily KM column
COL_KM_END    = 34  # AI — last daily KM column  (FIX: was 35, off by one)
COL_EXTRA_KM  = 35  # AJ — "Exception Kms in the Month"  (FIX: was 36)
COL_APPROVAL  = 36  # AK — "SM Approval"                 (FIX: was 37)

RATE_PER_KM   = 3
DATA_START_ROW = 6  # 1-based


class ExcelParser:

    def parse(self, file_bytes: bytes, filename: str, email_subject: str, sender_email: str) -> list[dict]:
        """
        Parse an Excel file from raw bytes.
        Returns a list of payroll record dicts.
        """
        records = []

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception as e:
            raise ValueError(f"Cannot open workbook: {e}")

        sheet = self._find_sheet(wb, filename)
        if sheet is None:
            raise ValueError(f"Could not find sheet 'Present absent' in {filename}")

        # Dynamically identify the daily KM columns
        # They usually have dates in Row 4 or "KM's" in Row 5
        row_4 = next(sheet.iter_rows(min_row=4, max_row=4, values_only=True), [])

        # NEW: derive period from actual datetime values in Row 4 (most reliable).
        # Falls back to filename/subject text if Row 4 has no datetime values.
        month_tag = self._extract_month_from_row4(row_4, filename, email_subject)
        row_5 = next(sheet.iter_rows(min_row=5, max_row=5, values_only=True), [])
        
        # ── Validate first 5 columns ──────────────────────────────────────────
        if len(row_5) < 5:
            log.warning(f"    Rejected {filename}: Has fewer than 5 columns in header row.")
            return []
            
        c0, c1, c2, c3, c4 = (str(c).lower() for c in row_5[:5])
        if ("employee id" not in c0 or 
            "employee" not in c1 or 
            "designation" not in c2 or 
            "ase manager" not in c3 or 
            "asm manager" not in c4):
            log.warning(f"    Rejected {filename}: First 5 columns do not match expected payroll format.")
            return []

        daily_col_indices = []
        max_cols = max(len(row_4), len(row_5))
        for col_idx in range(5, max_cols):
            val4 = row_4[col_idx] if col_idx < len(row_4) else None
            val5 = row_5[col_idx] if col_idx < len(row_5) else None
            
            is_date = val4 is not None
            is_km = str(val5).strip().lower() in ["km's", "kms", "km"]
            
            if is_date or is_km:
                daily_col_indices.append(col_idx)
                
        if not daily_col_indices:
            # Fallback if no dates/KM's found
            daily_col_indices = list(range(5, 35))

        skipped = 0
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
        ):
            if all(v is None for v in row):
                break

            record = self._parse_row(row, row_idx, filename, email_subject, month_tag, daily_col_indices, sender_email)
            if record:
                records.append(record)
            else:
                skipped += 1

        log.info(f"    Parsed {len(records)} valid rows, skipped {skipped} from {filename}")
        return records

    # ── Sheet finder ───────────────────────────────────────────────────────────

    def _find_sheet(self, wb, filename: str):
        target = "Present absent"
        if target in wb.sheetnames:
            return wb[target]
        for name in wb.sheetnames:
            if "present" in name.lower() or "absent" in name.lower():
                log.warning(f"Using sheet '{name}' as fallback in {filename}")
                return wb[name]
        log.warning(f"Sheet 'Present absent' not found in {filename}, using first sheet")
        return wb[wb.sheetnames[0]]

    # ── Row parser ─────────────────────────────────────────────────────────────

    def _parse_row(
        self, row: tuple, row_idx: int, filename: str,
        email_subject: str, month_tag: str, daily_col_indices: list[int],
        sender_email: str
    ) -> dict | None:

        emp_id   = self._safe_str(row, COL_EMP_ID)
        emp_name = self._safe_str(row, COL_EMP_NAME)
        if not emp_id or not emp_name:
            return None

        designation = self._safe_str(row, COL_DESIG)
        ase_manager = self._safe_str(row, COL_ASE)
        asm_manager = self._safe_str(row, COL_ASM)
        sm_approval = self._safe_str(row, COL_APPROVAL)

        # Sum dynamically identified daily KM columns
        daily_kms = []
        for col_idx in daily_col_indices:
            val = row[col_idx] if col_idx < len(row) else None
            daily_kms.append(self._to_number(val))

        total_extra_km = sum(daily_kms)
        amount_inr     = total_extra_km * RATE_PER_KM

        return {
            "Employee ID":    emp_id,
            "Employee Name":  emp_name,
            "Designation":    designation,
            "ASE Manager":    ase_manager,
            "ASM Manager":    asm_manager,
            "Month":          month_tag,
            "Total Extra KM": total_extra_km,
            "Amount INR":     amount_inr,
            "Source File":    filename,
            "Email Subject":  email_subject,
            "Sender Email":   sender_email,
            "Processed Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    # ── Helpers ────────────────────────────────────────────────────────────────

    def _safe_str(self, row: tuple, idx: int) -> str:
        if idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def _to_number(self, val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, TypeError):
            return 0.0

    def _extract_month_from_row4(self, row_4: tuple, filename: str, subject: str) -> str:
        """
        Primary period detection: reads actual datetime objects from Row 4.

        Row 4 contains the date for each daily KM column, e.g.:
          2026-03-21, 2026-03-22, ..., 2026-04-19
        The period tag is built from the first and last distinct months found:
          → "March - April"

        Falls back to _extract_month() (filename/subject text) if Row 4
        contains no datetime values (e.g. sheet format changed).
        """
        month_names = {
            1: "January", 2: "February", 3: "March",    4: "April",
            5: "May",     6: "June",     7: "July",      8: "August",
            9: "September", 10: "October", 11: "November", 12: "December",
        }

        # Collect all datetime values from row 4 (openpyxl returns them as datetime)
        dates = [
            v for v in row_4
            if isinstance(v, (datetime, date))
        ]

        if not dates:
            log.warning(
                f"No datetime values found in Row 4 of '{filename}' — "
                "falling back to filename/subject period detection."
            )
            return self._extract_month(filename, subject)

        # Sort to ensure correct order regardless of column arrangement
        dates_sorted = sorted(dates, key=lambda d: d if isinstance(d, date) else d.date())
        first = dates_sorted[0]
        last  = dates_sorted[-1]

        first_month = first.month if isinstance(first, date) else first.date().month
        last_month  = last.month  if isinstance(last,  date) else last.date().month

        if first_month == last_month:
            # All dates in one month — infer previous month as start
            m2_name = month_names[last_month]
            m1_num  = 12 if last_month == 1 else last_month - 1
            m1_name = month_names[m1_num]
            tag = f"{m1_name} - {m2_name}"
        else:
            tag = f"{month_names[first_month]} - {month_names[last_month]}"

        log.info(f"    Period detected from Row 4 dates: {tag} ('{filename}')")
        return tag

    def _extract_month(self, filename: str, subject: str) -> str:
        """
        Extract a period tag like 'March - April' from filename or email subject.
        """
        text = f"{filename} {subject}".lower()
        
        month_names = {
            "jan": "January", "feb": "February", "mar": "March", "apr": "April",
            "may": "May", "jun": "June", "jul": "July", "aug": "August",
            "sep": "September", "oct": "October", "nov": "November", "dec": "December",
        }
        month_order = list(month_names.keys())

        # Find all month mentions in the text
        matches = []
        for m in month_names.keys():
            for match in re.finditer(rf"\b{m}[a-z]*\b", text):
                matches.append((match.start(), m))
        
        matches.sort(key=lambda x: x[0])
        unique_months = []
        for _, m in matches:
            if m not in unique_months:
                unique_months.append(m)

        if len(unique_months) >= 2:
            m_a, m_b = unique_months[0], unique_months[1]
            idx_a = month_order.index(m_a)
            idx_b = month_order.index(m_b)
            
            # Sort adjacent months correctly
            if (idx_b - idx_a) % 12 == 1:
                m1, m2 = m_a, m_b
            elif (idx_a - idx_b) % 12 == 1:
                m1, m2 = m_b, m_a
            else:
                m1, m2 = (m_a, m_b) if idx_a < idx_b else (m_b, m_a)
                
            return f"{month_names[m1]} - {month_names[m2]}"
            
        elif len(unique_months) == 1:
            # Assume the mentioned month is the end of the payroll cycle
            m2 = unique_months[0]
            m2_idx = month_order.index(m2)
            m1_idx = (m2_idx - 1) % 12
            m1 = month_order[m1_idx]
            return f"{month_names[m1]} - {month_names[m2]}"
        
        # Fallback to current and next month
        m1_idx = datetime.now().month - 1
        m2_idx = (m1_idx + 1) % 12
        fallback = f"{month_names[month_order[m1_idx]]} - {month_names[month_order[m2_idx]]}"
        
        log.warning(f"Could not extract month from '{text}' — using {fallback}")
        return fallback